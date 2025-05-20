""""
Helper function to be used in the LassoRegBoostPipeline
"""
from pandas import DataFrame
from sklearn.model_selection import train_test_split
from typing import Union, Any, Dict, List, Tuple
import datetime
import pathlib
from tqdm import tqdm
import pandas as pd
import numpy as np
import shutil
import ast
from openpyxl import styles, load_workbook
from typing import Optional, Tuple
import warnings
from iterative_regression.effect_measures_plot import EffectMeasurePlot
import numpy.typing as npt
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score, confusion_matrix
from sklearn.linear_model import SGDClassifier, SGDRegressor
from sklearn.metrics import classification_report
from sklearn.model_selection import GridSearchCV, StratifiedKFold
from iterative_regression.iterative_regression import (IterativeRegression, remove_base_from_iteratives,
                                                       SelectNextBase)
from sklearn.model_selection import KFold, cross_val_score
from config.config import sections
import matplotlib.pyplot as plt
import json


def validate_input(input_dict: Dict[str, Any]) -> bool:
    """
    Validate the input configuration dictionary for required keys and value types.

    :param input_dict: Dictionary containing model configuration parameters.
    :return: bool, True if validation passes, otherwise raises ValueError.
    """
    required_keys = {
        'model_name': str,
        'path_features_from_base_register': (pathlib.Path, type(None)),
        'path_data': pathlib.Path,
        'path_output': pathlib.Path,
        'learning_task': str,
        'regression_target': (str, type(None)),
        'classification_num_class': (int, type(None)),  # Could be an integer or None
        'classification_target_class': (str, type(None)),
        'train': bool,
        'ml_algorithm': str,
        'split_gender': str,
        'TargetSelectionSplitting_discrete_column': str,
        'TargetSelectionSplitting_cont_column': str,
        'TargetSelectionSplitting_method': str,
        'TargetSelectionSplitting_stratify': str,
        'TargetSelectionSplitting_percent': float,
        'TargetSelectionSplitting_equal_samples_class': bool,
        'IterativeFeatureReg_output_path': str,
    }

    # Check for missing keys
    missing_keys = [key for key in required_keys if key not in input_dict.keys()]
    if missing_keys:
        raise ValueError(f"Missing configuration keys: {missing_keys}")

    # Check for type correctness
    type_errors = [
        f"Key '{key}' expects {expected_type}, but got {type(input_dict[key])} instead."
        for key, expected_type in required_keys.items()
        if not isinstance(input_dict[key], expected_type)
    ]

    if type_errors:
        raise ValueError("Type errors in configuration: " + " ".join(type_errors))

    # Additional checks
    class_num_class = input_dict['classification_num_class']
    class_target_class = input_dict['classification_target_class']
    # Check if paths are valid
    if input_dict['path_features_from_base_register'] is not None:
        if not input_dict['path_features_from_base_register'].exists():
            raise ValueError("Features file path does not exist.")

    if not input_dict['path_data'].exists():
        raise ValueError("Data file path does not exist.")

    if input_dict['learning_task'] not in ['regression', 'classification']:
        raise ValueError("Learning task must be classification or regression")

    if input_dict['learning_task'] == 'regression':
        if class_num_class is not None or class_target_class is not None:
            raise ValueError(f'When regression, the classification inputs must be none but got '
                             f'{class_num_class} and {class_target_class} ')
    else:
        if input_dict['regression_target'] is not None:
            raise ValueError('Regression target must be None when using the classification configuration')

    # Validation for classification settings
    if (class_num_class is None) != (class_target_class is None):
        raise ValueError(
            "Both 'classification_num_class' and 'classification_target_class' must be set to None or have values.")

    if input_dict['classification_num_class'] is not None and input_dict['classification_num_class'] < 2:
        raise ValueError("classification_num_class must be at least 2 or None.")

    # Validation on gender split
    if input_dict['split_gender'] not in ['male', 'female', 'both']:
        raise ValueError("split_gender must be 'male', 'female', or 'both'.")

    return True


def name_model_and_create_directory(model_path: pathlib.Path,
                                    input_dict: dict[str, any],
                                    name_model: bool = False) -> pathlib.Path:
    """
    Define the name of the model to later easily track the models folders.

    :param model_path: Path where the model directory will be created.
    :param input_dict: Dictionary containing model configuration parameters.
    :return: path of the results folder for the machine learning model
    """

    def format_string(s):
        # convert a string with format _a_b_c to ABC
        parts = s.split('_')
        formatted_parts = [parts[0]] + [part.capitalize() for part in parts[1:]]
        return ''.join(formatted_parts)

    now = datetime.datetime.now()
    # time_now = now.strftime("%y-%m-%d-%H-%M")
    time_now = now.strftime("%m-%d-%H-%M")
    if name_model:
        model_name_ml = ''
        # Append additional details from input_dict to the model name
        if input_dict.get('learning_task') == 'regression':
            model_name_ml = 'reg_'
        elif input_dict.get('learning_task') == 'classification':
            model_name_ml = 'class_'

        model_name_ml += time_now

        if 'ml_algorithm' in input_dict:
            model_name_ml += '_' + str(input_dict['ml_algorithm'])

        if input_dict.get('classification_num_class') is not None:
            model_name_ml += '_' + str(input_dict['classification_num_class'])

        if input_dict.get('classification_target_class') is not None:
            model_name_ml += '_' + str(input_dict['classification_target_class'])[:4]

        if input_dict.get('split_gender') is not None:
            model_name_ml += '_' + str(input_dict['split_gender'])
    else:
        model_name_ml = time_now + '_' + input_dict.get('model_name')

    model_name_ml = format_string(model_name_ml)
    # Create the full path for the model directory
    model_full_path = model_path.joinpath(model_name_ml)

    # Create the directory if it does not exist
    if not model_full_path.exists():
        model_full_path.mkdir(parents=True, exist_ok=True)
        print(f"Model path has been created with name: {model_full_path}")

    # Create a 'result' subdirectory inside the model directory
    # result_path = model_full_path.joinpath('result')
    # if not result_path.exists():
    #     result_path.mkdir(parents=True, exist_ok=True)
    #     print(f"Result path has been created with name: {result_path}")

    return model_full_path


def get_dataset(dataset_path: pathlib.Path,
                sections: dict,
                manual_col_removal: list[str] = None,
                columns_path: Union[pathlib.Path, None] = None,  # Make this optional by assigning `None` as default
                targets_name: Union[str, list[str]] = '', ) -> tuple[
    DataFrame, Union[DataFrame, list[str], None], Union[list[str], Any]]:
    """
    Read the dataset and if a columns path is given it will select. Multiple targets can be given
    :param dataset_path: pathlib.Path, path of the dataset
    :param columns_path: pathlib.Path, path where the list of the columns of interest are found
    :param manual_col_removal: list[str], columns manually identified for removal
    :param targets_name: Union[str, list[str]], targets we want to include
    :return: pd.DataFrame
    """
    pp_data = pd.read_csv(dataset_path, low_memory=False)
    # feature_names = ['']
    # columns = ['']

    if manual_col_removal is None:
        raise ValueError(f'Manual columsn to remove must be given at the input')

    # if columns_path is not None:
    #     # reed the features obtained from the regression model of feature selection saved as an excel columns
    #     # Check the file extension and read accordingly
    #     if columns_path.suffix == '.xlsx':
    #         feature_names = pd.read_excel(columns_path)
    #     elif columns_path.suffix == '.csv':
    #         feature_names = pd.read_csv(columns_path)
    #     else:
    #         str_warning = f'Features are not in the expected format (.csv, .xlsx)'
    #         warnings.warn(str_warning)
    #     feature_names = feature_names.iloc[:, -1].tolist()
    #     if feature_names is not None:
    #         # Convert the DataFrame of feature names to a list, coming from base register so is the last column
    #         columns = feature_names.copy()  # Assuming the names are in the first column
    #         # Handle the inclusion of targets
    #         if isinstance(targets_name, str) and targets_name:
    #             columns.append(targets_name)
    #         elif isinstance(targets_name, list) and targets_name:
    #             columns.extend(targets_name)
    # elif sections is not None:

    # # get all the features, method designed specific to the Bioserenity dataset
    # post_sleep_section = [sec for sec in sections if sec == 'PostSleep_']
    # feature_sections = pp_data.columns[pp_data.columns.str.startswith(tuple(sections))].to_list()

    lab_columns = [
        "ID",
        "DOB_YEAR",
        "Study_YEAR",
        "LAB_CAT",
        "MD_Identifyer",

        "BIPAP",
        "Oxygen",
        "RDI",
        "LowSat",
        "TIB",
        "TST",
        "SME",
        "SO",
        "ROL",
        "AI",
        "PLMS",
        # "DI",  # TODO: use the ODI (DI), this is at 3%
        "SEN",
        "SAO2_Per",
        "LPS",

        "ISL",
        "USL",
        "WASO",

        "Latency_Stage1",
        "Latency_Stage2",
        "Latency_Stage3",
        "Latency_SWS",
        "Latency_REM",
        "Latency_NREM",

        "Duration_Stage1",
        "Duration_Stage2",
        "Duration_Stage3",
        "Duration_Stage4",
        "Duration_SWS",
        "Duration_REM",
        "Duration_NREM",
        "Duration_MVT",

        "Percentage_Stage1",
        "Percentage_Stage2",
        "Percentage_Stage3",
        "Percentage_Stage4",
        "Percentage_SWS",
        "Percentage_REM",
        "Percentage_NREM",
        "Percentage_MVT",

        "BMI_Square",
        # "Age_raw",
        "BMI_raw",
        "BMI_Square_raw",

        "AI_REM",
        "AI_NREM",
        "AHI_REM",
        "AHI_NREM",
        "RDI_REM",
        "RDI_NREM",
        "BMI_Square",
        "AI_log1p",
        'AI_multiclass',

        # "AHI_multiclass",
        # "AHI_log1p",
        "AI_REM_multiclass",
        "AI_REM_log1p",
        "AI_NREM_multiclass",
        "AI_NREM_log1p",
        "AHI_REM_multiclass",
        "AHI_REM_log1p",
        "AHI_NREM_multiclass",
        "AHI_NREM_log1p",
        "SAO2_Per_log1p",
        "LowSat_log1p",]
    # manual columns removal
    post_sleep_columns = pp_data.columns[pp_data.columns.str.startswith('PostSleep_')].to_list()
    columns_to_remove = [col for col in manual_col_removal if col in pp_data.columns]
    columns_to_remove = columns_to_remove + lab_columns + post_sleep_columns
    pp_data.drop(columns=columns_to_remove, inplace=True)

    duplicate_columns = pp_data.columns[pp_data.columns.duplicated()].unique()

    # Remove columns from feature_names
    feature_names = [col for col in pp_data.columns if 'AHI' not in col]  # remove the target variable from the features
    columns = [*pp_data.columns]

    return pp_data, feature_names, columns


def slice_by_gender(df: pd.DataFrame,
                    gender_column: str,
                    gender_select: str):
    """
    split the dataset by gender and drop the gender feature if not both are used
    :param df: dataset
    :param gender_column: str, gender column in the dataset
    :param gender_select: gender to select
    :return:
    """
    gender_code = {'male': 1, 'female': 0}
    if gender_select == 'both':
        print(f'Gender Selection: both genders (all dataset) - {df.shape}')
        return df
    elif gender_select == 'male':
        df = df.loc[df[gender_column] == gender_code.get('male')]
        df.drop(columns=[gender_column], inplace=True)
        print(f'Gender Selection: males - {df.shape}')
        return df
    else:
        df = df.loc[df[gender_column] == gender_code.get('female')]
        df.drop(columns=[gender_column], inplace=True)
        print(f'Gender Selection: females - {df.shape}')
        return df
    pass


def reduction_elastic_net_lasso(frame: Union[pd.DataFrame, dict[str, pd.DataFrame]],
                                target: str,
                                model: str,
                                feature_names: list[str],
                                output_path: pathlib.Path,
                                sections: Optional[dict] = None,
                                stratify_column: Optional[str] = None,
                                ) -> tuple[list[str], DataFrame]:
    """
    Perform lasso for classification or linear regression problem and select significant features that are higher than
    a certain threshold in absolute value.

    If train/val splits are received, the lasso is trained/tested on these splits. Otherwise, it opens the
    Bioserenity dataset with all the desired features.

    LassoClassification outputs the feature coefficients for each class separately. Therefore, to select them
    we select those that in at least one class satisfies the threshold in absolute value.

    :param frame:Union[pd.DataFrame, dict[str, pd.DataFrame]], frame or dict of splits {train: , val: }
    :param target: str, target to run the lasso model
    :param model: str, if to regression or classification
    :param sections: dict, containing the sections we would like to remove if all the frame is given
    :param feature_names: list[str], is given, subset of features to run the lasso
    :param output_path: pathlib.path, path to save the csv
    :param stratify_column: str, column to stratify the dataset if the frame is given
    :return:
            list[str] significant feature names given by lasso
    """
    GRID_SEARCH = False

    if not model in ['regression', 'classification']:
        raise ValueError(f'model must be either classification or regression')

    if isinstance(frame, dict):
        # if the splits are given use them
        train_df = frame.get('train')
        test_df = frame.get('val')
    else:
        # else make them from the frame
        if feature_names is None and sections is not None:
            sections = [sec for sec in sections if sec != 'PostSleep_']
            feature_names = frame.columns[frame.columns.str.startswith(tuple(sections))].to_list()

        train_df, test_df = train_test_split(frame,
                                             test_size=0.2,
                                             stratify=frame[stratify_column],
                                             random_state=0)

    if not output_path.exists():
        output_path.mkdir(parents=True,
                          exist_ok=True)

    # remove target from the columns
    col_features = feature_names.copy()

    train_x = train_df[col_features]
    train_y = train_df[target]

    test_x = test_df[col_features]
    test_y = test_df[target]

    if model == 'classification':
        elastic_net_model = SGDClassifier(loss='log_loss',
                                          penalty='elasticnet',
                                          l1_ratio=0.5,
                                          alpha=0.001,
                                          max_iter=2000,
                                          tol=None,
                                          random_state=42,
                                          learning_rate='adaptive',
                                          eta0=0.01,
                                          warm_start=True)
    else:
        elastic_net_model = SGDRegressor(loss='squared_error',
                                         penalty='elasticnet',
                                         l1_ratio=0.5,
                                         alpha=0.001,
                                         max_iter=2000,
                                         tol=None,
                                         random_state=42,
                                         learning_rate='adaptive',
                                         eta0=0.01,
                                         warm_start=True)
    if GRID_SEARCH:
        # GridSearchCV
        param_grid = {
            'alpha': [0.001, 0.01, 0.1, 1, 2],  # Regularization strength
            'l1_ratio': [0.1, 0.5, 0.9],  # Balance between L1 and L2 regularization
        }
        cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        if model == 'classification':
            grid_search = GridSearchCV(elastic_net_model,
                                       param_grid,
                                       cv=cv,
                                       scoring='accuracy',
                                       verbose=1,
                                       n_jobs=-1)
            grid_search.fit(train_x, train_y)
            # Predict on test set with the best model
            best_model = grid_search.best_estimator_
            pred_test = best_model.predict(test_x)

            # grid_search.best_params_
            # Out[3]: {'alpha': 0.001, 'l1_ratio': 0.5}
            #
            # Classification report
            print(classification_report(test_y, pred_test))
            report_df = pd.DataFrame(classification_report(test_y, pred_test, output_dict=True)).transpose()
            print(report_df)
            report_df.to_excel(f'SGDClassifier_report.xlsx')
        else:
            # neg_mean_squared_error : closer to the mean
            # neg_mean_absolute_error: closer to the median , useful significant outliers don’t want to influence model
            best_params = None
            best_score = -np.inf  # Negative because we use neg_mean_squared_error

            # Use KFold cross-validation
            kf = KFold(n_splits=5, shuffle=True, random_state=42)

            # Iterate over the parameter grid
            for alpha in param_grid['alpha']:
                # Create a Lasso model with the current alpha

                # Perform cross-validation
                scores = cross_val_score(elastic_net_model,
                                         train_x,
                                         train_y,
                                         cv=kf,
                                         scoring='neg_mean_squared_error',
                                         n_jobs=-1)

                # Calculate the mean score
                mean_score = np.mean(scores)

                # If the current score is better than the best score, update the best parameters and score
                if mean_score > best_score:
                    best_score = mean_score
                    best_params = {'alpha': alpha}
            # Output the best parameters and score
            print("Best parameters found: ", best_params)
            print("Best cross-validation score (negative MSE): ", best_score)

    # Fit the model
    elastic_net_model.fit(train_x, train_y)

    # Predict on training and validation set
    pred_train = elastic_net_model.predict(train_x)
    pred_test = elastic_net_model.predict(test_x)
    if model == 'classification':
        print(classification_report(test_y, pred_test))
        report = classification_report(test_y, pred_test, output_dict=True)
        report_df = pd.DataFrame(report).transpose()
        conf_matrix = confusion_matrix(y_true=test_y, y_pred=pred_test)
        conf_matrix_df = pd.DataFrame(conf_matrix,
                                      columns=['Predicted Class 0',
                                               'Predicted Class 1',
                                               'Predicted Class 2',
                                               'Predicted Class 3'],
                                      index=['Actual Class 0',
                                             'Actual Class 1',
                                             'Actual Class 2',
                                             'Actual Class 3'])

        report_df.to_csv(output_path.joinpath('lasso_reduction_metrics_val_set.csv'),
                         index=True)
        conf_matrix_df.to_csv(output_path.joinpath('lasso_reduction_confmatrix_val_set.csv'),
                              index=True)

        report = classification_report(train_y, pred_train, output_dict=True)
        report_df = pd.DataFrame(report).transpose()
        conf_matrix = confusion_matrix(y_true=train_y, y_pred=pred_train)
        conf_matrix_df = pd.DataFrame(conf_matrix,
                                      columns=['Predicted Class 0',
                                               'Predicted Class 1',
                                               'Predicted Class 2',
                                               'Predicted Class 3'],
                                      index=['Actual Class 0',
                                             'Actual Class 1',
                                             'Actual Class 2',
                                             'Actual Class 3'])

        report_df.to_csv(output_path.joinpath('lasso_reduction_metrics_train_set.csv'),
                         index=False)
        conf_matrix_df.to_csv(output_path.joinpath('lasso_reduction_confmatrix_train_set.csv'),
                              index=False)

        # now we need to select the most important features with lasso
        feature_importance = elastic_net_model.coef_
        lasso_coeff_df = pd.DataFrame(np.nan,
                                      columns=[f'class_{idx}' for idx in range(0, len(feature_importance))],
                                      index=[*train_x.columns])
        for idx_coeff in range(0, len(feature_importance)):
            lasso_coeff_df.iloc[:, idx_coeff] = feature_importance[idx_coeff]
    else:
        def evaluate_regression(y_pred: npt.NDArray, y_true: npt.NDArray) -> dict:
            return {
                "RMSE": np.sqrt(mean_squared_error(y_true, y_pred)),
                "MAE": mean_absolute_error(y_true, y_pred),
                "R2": r2_score(y_true, y_pred),
            }

        report = evaluate_regression(y_pred=pred_test, y_true=test_y)
        report_df = pd.DataFrame.from_dict(report, orient='index', columns=['Values'])
        report_df.to_csv(output_path.joinpath('lasso_reduction_metrics_val_set.csv'))

        report = evaluate_regression(y_pred=pred_train, y_true=train_y)
        report_df = pd.DataFrame.from_dict(report, orient='index', columns=['Values'])
        report_df.to_csv(output_path.joinpath('lasso_reduction_metrics_train_set.csv'))

        lasso_coeff_df = pd.DataFrame({'Feature': [*train_x.columns],
                                       'Importance': elastic_net_model.coef_})
    lasso_coeff_df.to_csv(output_path.joinpath('lasso_reduction_coeff_list.csv'),
                          index=True)
    # get the rows where at least one of the columns is higher than 0.09 in absolute value
    filtered_df = lasso_coeff_df[(lasso_coeff_df.abs() > 0.09).any(axis=1)]
    filtered_df.to_csv(output_path.joinpath('lasso_reduction_coeff_list_filtered.csv'),
                       index=True)

    return filtered_df.index.to_list(), lasso_coeff_df


def remove_base_from_iteratives(base_features: list,
                                iterative_features: list) -> list:
    """
    Remove base features from the iterative features list.

    Args:
        base_features (list): List of base features.
        iterative_features (list): List of iterative features.

    Returns:
        list: List of iterative features with base features removed.
    """
    return [iter_feat for iter_feat in iterative_features if not iter_feat in base_features]


def max_percentage(n_elements: int,
                   percent: int = 30) -> int:
    return int(np.floor((percent / 100) * n_elements))


def rename_and_save_table_files(result_folder_path: pathlib.Path,
                                name_table_file: str,
                                output_path: pathlib.Path):
    """
    To present the results without a nested structure of folder and file, this function will pass all the tabels of
    interest into a single folder and rename the table files with the folder name they were originated.
    :param result_folder_path:
    :param name_table_file:
    :param output_path:
    :return:
    """
    # Check if the output folder exists, and create it if not
    output_path.mkdir(parents=True, exist_ok=True)
    folders = list(result_folder_path.iterdir())
    for folder in tqdm(folders, desc='Moving and renaming files', unit='folder'):
        if folder.is_dir() and folder.name.split('_')[0].isnumeric():
            old_file_path = folder / name_table_file
            if old_file_path.is_file():
                new_file_name = str(folder.name) + f".{name_table_file.split('.')[1]}"
                new_file_path = output_path / new_file_name
                # Check if new file already exists and remove it if it does
                if new_file_path.is_file():
                    new_file_path.unlink()

                shutil.copy(old_file_path, new_file_path)


def check_all_in_b(list_a: list, list_b: list):
    """
    Check if all the elements of list a are in list b
    :param list_a:
    :param list_b:
    :return:
    """
    return all(item in list_b for item in list_a)


def create_pretty_table(table_path: pathlib.Path,
                        output_path: Optional[pathlib.Path] = None):
    """
    From the single_coefficient_summary.xlsx table, with columns
        ['variable', 'P>|t|', 'No. Observations', 'odds', 'odds_ci_low', 'odds_ci_high', 'responses_count']
    we will modify the xlsx so it can be presented better reducing the column, using multiple strings in a single
    row and centering.


    :param table_path:
    :param output_path: path of the output, if None then it overwrite the file in table_path
    :return:
    """

    def sort_dict_by_keys(input_dict: dict) -> dict:
        """Check if all keys are either float or int, then sort them"""
        # check if all keys are either int or float
        if all(isinstance(k, (int, float)) for k in input_dict.keys()):
            # sort the dictionary
            sorted_dict = dict(sorted(input_dict.items()))
            return sorted_dict
        else:
            return input_dict

    def dict_to_string(x: dict) -> str:
        """
        for presentation in Excel convert eh dictionary to string of multiple rows

        Example:
        response_count = {4.0: (47072, 34.8), 3.0: (31859, 23.5), 2.0: (22449, 16.6), 1.0: (17062, 12.6),
                  0.0: (16961, 12.5)}
        results in
            4: (47072, 34.8)
            3: (31859, 23.5)
            2: (22449, 16.6)
            1: (17062, 12.6)
            0: (16961, 12.5)
        :param x:
        :return:
        """
        x_sorted = sort_dict_by_keys(input_dict=x)
        return '\n'.join(f'{key}: {value} \n' for key, value in x_sorted.items())

    def check_columns(df: pd.DataFrame, required_columns: list[str]) -> bool:
        return all(column in df.columns for column in required_columns)

    if table_path.suffix != '.xlsx':
        raise ValueError(f'Only xlsx format are acceptable got {table_path.suffix}')

    if output_path is None:
        output_path = table_path

    # get only the columns of interest
    col_interest = ['variable', 'P>|t|', 'No. Observations', 'odds', 'odds_ci_low', 'odds_ci_high',
                    'responses_count']
    df = pd.read_excel(table_path)

    if not check_columns(df=df, required_columns=col_interest):
        str_warn = f'Unable to make column for {table_path.name}, required columns not found'
        warnings.warn(str_warn)
        return None

    df = df[col_interest]

    # the multiple response count parse it a string so we can have the multi row in the final xlsx
    df['responses_count'] = df['responses_count'].apply(ast.literal_eval)
    df['Response (Count, Percent)'] = df['responses_count'].apply(dict_to_string)

    # combine the odds into a single column
    df['OR'] = df.apply(lambda row: f"{row['odds']} ({row['odds_ci_low']},{row['odds_ci_high']})", axis=1)
    df.drop(columns=['odds', 'odds_ci_low', 'odds_ci_high', 'responses_count'],
            inplace=True)

    df.rename(columns={
        'P>|t|': 'p-value',
    },
        inplace=True)

    # re-order the columns
    df = df[['variable', 'p-value', 'OR', 'No. Observations', 'Response (Count, Percent)']]

    # temporarly save the file
    df.to_excel('tmp_wrap_text.xlsx',
                index=False,
                sheet_name='Sheet1')

    # Make the \n applicable in excel and aling the cells as we want for presentation
    workbook = load_workbook('tmp_wrap_text.xlsx')
    worksheet = workbook['Sheet1']

    # Iterate through each row and cell to apply style alignment
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = styles.Alignment(wrapText=True,
                                              horizontal='center',
                                              vertical='center')

    # Iterate over the columns so they can have the column width adjustment
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # save the pretty table
    print(f'Pretty table applied to: {output_path.name}\n')
    workbook.save(output_path)


def plot_or_ci(pretty_table_path: pathlib.Path,
               figsize: Tuple[int, int] = (14, 16),
               show: Optional[bool] = False):
    """
    From the generated pretty tables xlsx file, were we have an excel with the structure:
    Index(['variable', 'p-value', 'OR', 'No. Observations', 'Response (Count, Percent)'],dtype='object')

                   variable  ...                          Response (Count, Percent)
        0  SA_Sleep_Stop_Breath  ...  0.0: (61899, 47.9) \n\n1.0: (17720, 13.7) \n\n...
        1        SA_Sleep_Snore  ...  0.0: (16961, 12.5) \n\n1.0: (17062, 12.6) \n\n...

    We will generate a traditional Epidemiology plot to show the OR, CI of each variable and include a vertical
    line were the level of significance is located with alpha = 0.05 with Bonferroni correction.

    :param pretty_table_path:pathlib.Path, path of the xlsx file
    :param figsize: Optional[tuple], size of the figure
    :return:
    """
    # pretty_table_path = pathlib.Path(r'C:\Users\giorg\OneDrive - Fundacion Raices Italo Colombianas\projects\miglab_bioserenity\results\presentation\reduced_model_53_features\6_PH3_Before_Bed_Sleep_Aids_iter_reg.xlsx')
    output_path = pretty_table_path.with_suffix('.png')
    df_table = pd.read_excel(pretty_table_path)
    # unstack the OR and CI from single cell to multiple columns
    df_table[['OR', 'ci_low_bound', 'ci_high_bound']] = df_table['OR'].str.extract(r'([\d.]+)\s+\(([\d.]+),([\d.]+)\)')
    columns = ['variable', 'OR', 'ci_low_bound', 'ci_high_bound', 'p-value']
    df_table = df_table[columns]
    df_table[columns[1::]] = df_table[columns[1::]].astype(float)
    df_table.sort_values(by='p-value',
                         inplace=True)
    alpha_corr = 0.05 / df_table.shape[0]
    forest_plot = EffectMeasurePlot(label=df_table.variable.tolist(),
                                    effect_measure=df_table.OR.tolist(),
                                    lcl=df_table.ci_low_bound.tolist(),
                                    ucl=df_table.ci_high_bound.tolist(),
                                    p_value=df_table['p-value'].tolist(),
                                    alpha=alpha_corr)

    forest_plot.plot(figsize=figsize,
                     path_save=output_path,
                     show=show)
    plt.clf()
    plt.cla()
    plt.close()


def pad_list(input_list, target_len, pad_value=None) -> list:
    """Pad a list with nans so it has the target length"""
    return input_list + [pad_value] * (target_len - len(input_list))


def iterative_feature_selection(frame: pd.DataFrame,
                                features_from_lasso: list[str],
                                col_manual_removal: list[str],
                                model_name: str,
                                output_path: pathlib.Path,
                                target_continuous: str = 'AHI_log1p',
                                max_iterations: Optional[int] = 100,
                                ) -> list[str]:
    """
    Function wrapper for the iterative linear regression algorithm. It starts with a set of base features
    found in the dataset and then uses the inputs features_from_lasso which are the selected features from lasso
    to iterate and test for significance using linear regression, OR, and CI.

    The objective of the function is to further reduce the features from col_manual_removal and output the
    reduced list.

    Results of the class will be saved in the output path which is expected to be the path of the model

    :param frame: pd.DataFrame,
    :param features_from_lasso: list[str],
    :param col_manual_removal: list[str],
    :param model_name: str,
    :param output_path:pathlib.Path,
    :param target_continuous: AHI_log1p,
    :param max_iterations: int,
    :return:
        list[str], Further reduced features returned as a list
    """
    # add the model folder the linear regression feature results
    if not output_path.exists():
        output_path.mkdir(parents=True,
                          exist_ok=True)
    # Presentation folder to have the tables in a presentable and clear format
    # presentation_path = output_path.joinpath('presentation', model_name)
    presentation_path = output_path.joinpath('publish')
    if not presentation_path.exists():
        presentation_path.mkdir(parents=True,
                                exist_ok=True)

    base_logs_path = output_path.joinpath(f'BaseLogs')
    if not base_logs_path.exists():
        base_logs_path.mkdir(parents=True,
                             exist_ok=True)
    # base features
    base_features = [
        'DisplayGender',
        'BMI',
        'Age',
        'Race_1',
        'Race_2',
        # 'Race_3',
        'Race_4',
        'Race_5',
        # 'Race_6',
        'SA_Sleep_Snore',
        'SA_Sleep_Stop_Breath',
        'MH_High_Blood_Pressure',
        'SA_Sleep_Restless',
        'PH2_Head_Ache',
        'Respiratory_Health_Status',
        'PH2_Dentures'
    ]
    # remove the manually selected features where we used correlation and medical criteria
    base_features = [base_feat for base_feat in base_features if base_feat not in col_manual_removal]
    # remove base features that are not in frame e.g., when using a single gender and DisplayGender is not present
    base_features = [base_feat for base_feat in base_features if base_feat in frame.columns]

    features_from_lasso = [lasso_feat for lasso_feat in features_from_lasso if not lasso_feat in col_manual_removal]

    iterative_features = remove_base_from_iteratives(base_features=base_features,
                                                     iterative_features=features_from_lasso)
    # cluster all the columns that we will use, including target
    columns = base_features.copy()
    columns.extend(iterative_features)
    columns.append(target_continuous)
    # from collections import Counter
    # column_counts = Counter(columns)
    # duplicate_columns = [column for column, count in column_counts.items() if count > 1]
    columns = list(set(columns))
    frame = frame[columns]

    max_len = max(len(base_features), len(iterative_features))
    report_df = pd.DataFrame({'Initial Model': pad_list(base_features, max_len),
                              'Features Iterate': pad_list(iterative_features, max_len),
                              'sections': pad_list(sections, max_len)})

    if target_continuous not in frame.columns:
        raise ValueError(f'Target {target_continuous} is not in the columns of the data\n')

    select_next_base = SelectNextBase(
        criteria='P>|t|',
        column_variables='variable',
        alpha=0.05,
        output_path=base_logs_path.joinpath(f'{model_name}_trial.xlsx'),
        # to_ignore_significance=significants_to_ignore
    )

    next_base = base_features
    iterative_count = 0
    len_iter_feat = len(iterative_features)
    iter_limit = max_percentage(n_elements=len_iter_feat,
                                percent=30)
    print(f'While Loop will be terminated when the iterative features ({len_iter_feat}) reaches a minimum of '
          f'{iter_limit} or when the next base is emtpy, starting with {len(next_base)} candidates ')
    # all conditions must be true for the loop to continue
    while len_iter_feat != iter_limit and len(next_base) > 0 and (
            iterative_count <= max_iterations):  # termination condition
        print(f'Iterative Count {iterative_count}\n')
        # iterative_features = remove_base_from_iteratives(base_features=next_base,
        #                                                  iterative_features=iterative_features)
        iterativeregression = IterativeRegression(
            data=frame,
            base_features=next_base,
            iterative_features=iterative_features,
            target=target_continuous,
            out_path=output_path.joinpath('iter'),
            trial_name=f'{iterative_count}_{next_base[-1]}',
        )

        iterativeregression.fit_iterative_models()

        model_results = iterativeregression.get_iteration_results()

        next_base: list = select_next_base.select_base(iterations_result=model_results,
                                                       current_base=iterativeregression.base_features)

        next_iterative_features: list = remove_base_from_iteratives(base_features=next_base,
                                                                    iterative_features=iterative_features)

        iterative_features = next_iterative_features.copy()
        iterative_count += 1

    # save the history of base selection
    base_register = select_next_base.save_base()
    if select_next_base.to_ignore is not None:
        select_next_base.save_to_ignore_count()

    base_register.to_excel(presentation_path.joinpath('BaseReg.xlsx'),
                           index=False)

    rename_and_save_table_files(
        result_folder_path=output_path.joinpath('iter'),
        name_table_file='SingleCoeffSum.xlsx',
        output_path=presentation_path,
    )

    for file_name in presentation_path.glob(pattern='*.xlsx'):
        # print(file_name
        create_pretty_table(table_path=file_name)

    report_df.to_excel(presentation_path.joinpath('InitialVsIterativeList.xlsx'),
                       index=False)

    #  Epidemiology plots of OR and CI
    for table_path_ in presentation_path.glob('*.xlsx'):
        file_name = str(table_path_.name)
        print(file_name)
        if "_" in file_name:
            if file_name.split('_')[0].isdigit():
                plot_or_ci(pretty_table_path=table_path_,
                           figsize=(14, 16),
                           show=False)
    # TODO: return the OR and CI of the last model (significant one) so we can later compare them
    col_significant_features = base_register.iloc[:, -1].tolist()
    return col_significant_features


def generate_output_report(
        features_initial: list[str],
        features_after_lasso: list[str],
        features_after_reg: list[str],
        metrics_ml: pd.DataFrame,
        lasso_coeff_df: pd.DataFrame,
        output_path: pathlib.Path,
        input_config: dict,
        splits_target: dict[str, pd.Series]) -> dict:
    """
    Collect the main outputs and feature results in a single dataframe, so that we can easily compare across models
    :param features_initial: list[str], initial features to input in the pipline
    :param features_after_lasso: list[str], features after lasso reduction
    :param features_after_reg: list[str], features after linear regression reduction
    :param metrics_ml: pd.DataFrame, frame with the metrics of each split
    :param output_path: pathlib, path to save the output as a .csv
    :param input_config: dict, input configuration dictionary to get some parameters
    :return:
    - dict of the summary output
    """
    if not output_path.exists():
        output_path.mkdir(parents=True)

    if input_config.get('learning_task') == 'classification':
        metrics = 'F1 Score'
    else:
        metrics = 'RMSE'

    metrics_ml.set_index('Splits', inplace=True)

    output_dict = {
        'model_name': input_config.get('model_name'),
        'model_path': str(output_path),
        'learning_task': input_config.get('learning_task'),
        'split_gender': input_config.get('split_gender'),
        'features_initial': str(features_initial),
        'features_afterlasso': str(features_after_lasso),
        'features_afterreg': str(features_after_reg),
        'n_features_initial': len(features_initial),
        'n_features_afterlasso': len(features_after_lasso),
        'n_features_afterreg': len(features_after_reg),
        f'train_{metrics}': metrics_ml.at['train', metrics].round(3),
        f'val_{metrics}': metrics_ml.at['valid', metrics].round(3),
        f'test_{metrics}': metrics_ml.at['test', metrics].round(3),
        'TargetSelectionSplitting_method': input_config.get('TargetSelectionSplitting_method'),
        'TargetSelectionSplitting_percent': None,
        'TargetSelectionSplitting_equal_samples_class': input_config.get(
            'TargetSelectionSplitting_equal_samples_class'),
    }

    if lasso_coeff_df is not None:
        lasso_coeff_json = lasso_coeff_df.to_json()
        output_dict['lasso_coefficients'] = lasso_coeff_json

    if input_config.get('TargetSelectionSplitting_method') in ['boundaries', 'interior_domain']:
        output_dict['TargetSelectionSplitting_percent'] = input_config.get('TargetSelectionSplitting_percent')

    splits_target = {key: val.to_list() for key, val in splits_target.items()}
    output_dict['splits_target'] = splits_target

    with open(output_path.joinpath('output_report.json'), 'w') as f:
        json.dump(output_dict, f)

    return output_dict


def get_folder_with_highest_number(parent_path: Union[str, pathlib.Path],
                                   filename: str = "SingCoeffSumm.xlsx") -> Union[None, pathlib.Path]:
    """
    get the last OR table made by the regression model.
    The folder have a numerica hierarchy e.g., 0_, 1_, 2_. It uses the numers to get the last folder.
    In the last folder, it searches for the filename and returns it
    :param parent_path: parent path of the subfolders
    :param filename: str, file name and preffix we are searching in the folders
    :return: path of the file
    """
    if isinstance(parent_path, str):
        parent_path = pathlib.Path(parent_path)

    highest_number = -1
    highest_folder = None

    for folder in parent_path.iterdir():
        if folder.is_dir():
            try:
                folder_number = int(folder.name.split('_')[0])  # Extract the number before the first underscore
                if folder_number > highest_number:
                    highest_number = folder_number
                    highest_folder = folder
            except ValueError:
                # Skip folders that don't follow the naming convention
                continue

    if highest_folder:
        excel_file = highest_folder.joinpath(filename)
        if excel_file.exists():
            return excel_file
    return None